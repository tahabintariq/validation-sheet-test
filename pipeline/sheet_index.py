"""
Reads "Qtrly Results Updated" and builds the lookup the pipeline needs:

  - which Tadawul codes are in the sheet, and on which row(s)
  - which sector section each row sits in
  - which codes are out of scope

Sections are read from the sheet itself every run - a section header is a row
with something in column A but no Tadawul code. So if someone moves a company
between sectors, or adds a sector, the pipeline follows automatically instead
of relying on a list baked into code.

Out of scope:
  - everything in the "Insurance" section (client's instruction)
  - Arabian Shield (8070), an insurer that sits in the Small-Cap section
  - NOT Rasan (8313), which is insurance technology, not an insurer, and is
    tracked under Information Technology

Some companies appear twice (e.g. Aldrees 4200 in both Energy and Retail), so a
code maps to a LIST of rows and figures are written to all of them.
"""

import re
import sys
from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parent.parent / "Saudi Valuation Sheet AI.xlsx"
SHEET = "Qtrly Results Updated"

# Row 2 holds the metric name, row 3 the quarter label. Everything else is found
# by reading those headers rather than by counting columns: if someone inserts a
# column, positions shift but names do not. A positional read would keep working
# silently and match the wrong company - the worst kind of failure here.
HEADER_ROW_METRIC = 2
HEADER_ROW_QUARTER = 3
CODE_HEADER = "tadawul code"
METRICS = ["Net Income", "Revenue", "Gross Profit", "Operating Profit"]
NOMU_HALFYEAR_SECTION = "NOMU Halfyear Companies"

SKIP_SECTIONS = {"insurance"}
SKIP_CODES = {"8070"}          # Arabian Shield - insurer filed under Small-Cap
KEEP_CODES = {"8313"}          # Rasan - insurtech, stays in scope


def normalise_half(label):
    """'1H2026' -> (2026, 1). None if not a half-year label."""
    if label is None:
        return None
    m = re.fullmatch(r"([12])H(\d{4})", str(label).strip().upper().replace(" ", ""))
    return (int(m.group(2)), int(m.group(1))) if m else None


def find_half_year_header_rows(all_rows, min_labels=4):
    """Every row that relabels the columns as 1H/2H periods.

    There is more than one: row 377 serves the Nomu half-year section (and the
    IPO half-yearly section below it, which has no header of its own), and row
    212 serves the REIT section. A row's labels are whichever half-year header
    sits closest above it, so the wrong section's labels can never be used.
    """
    return [i for i, row in enumerate(all_rows, start=1)
            if sum(1 for v in row if normalise_half(v)) >= min_labels]


def normalise_quarter(label):
    """'1Q2026' and '2026Q1' both -> (2026, 1). Returns None if not a quarter label.

    The sheet uses both spellings - the Net Income block says 1Q2026, the Revenue
    block says 2026Q1 - so neither form can be assumed.
    """
    if label is None:
        return None
    s = str(label).strip().upper().replace(" ", "")
    m = re.fullmatch(r"([1-4])Q(\d{4})", s)
    if m:
        return int(m.group(2)), int(m.group(1))
    m = re.fullmatch(r"(\d{4})Q([1-4])", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def load_index(workbook=WORKBOOK, sheet=SHEET):
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"FATAL: sheet '{sheet}' not found. Tabs present: {wb.sheetnames}")
    ws = wb[sheet]

    rows_iter = ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)
    all_rows = list(rows_iter)
    metric_hdr = list(all_rows[HEADER_ROW_METRIC - 1])
    quarter_hdr = list(all_rows[HEADER_ROW_QUARTER - 1])

    # --- locate the Tadawul code column BY NAME, never by position -----------
    code_col = None
    for i, h in enumerate(metric_hdr):
        if h and str(h).strip().lower() == CODE_HEADER:
            code_col = i
            break
    if code_col is None:
        raise SystemExit(
            f"FATAL: no '{CODE_HEADER}' header found in row {HEADER_ROW_METRIC} of '{sheet}'.\n"
            "  The sheet layout has changed. Refusing to guess which column holds the code -\n"
            "  guessing would match figures to the wrong companies."
        )

    # --- locate each metric's columns, by name and as one contiguous block ---
    # A metric block is a RUN of adjacent columns sharing the same row-2 name.
    # Taking the longest run excludes strays that carry the same name far away
    # (there is one: a lone 'Net Income / 4Q2024' at EG), which a plain name
    # search would happily return.
    half_header_rows = find_half_year_header_rows(all_rows)

    metric_cols, metric_half_cols, metric_blocks = {}, {}, {}
    for m in METRICS:
        hits = [i for i, h in enumerate(metric_hdr) if h and str(h).strip() == m]
        runs, run = [], []
        for i in hits:
            if run and i == run[-1] + 1:
                run.append(i)
            else:
                if run:
                    runs.append(run)
                run = [i]
        if run:
            runs.append(run)
        block = max(runs, key=len) if runs else []
        metric_blocks[m] = (block[0] + 1, block[-1] + 1) if block else None

        cols = {}
        for i in block:
            q = normalise_quarter(quarter_hdr[i] if i < len(quarter_hdr) else None)
            if q and q not in cols:
                cols[q] = i + 1
        metric_cols[m] = cols

        per_header = {}
        for hr in half_header_rows:
            hdr = list(all_rows[hr - 1])
            halves = {}
            for i in block:
                h = normalise_half(hdr[i] if i < len(hdr) else None)
                if h and h not in halves:
                    halves[h] = i + 1
            per_header[hr] = halves
        metric_half_cols[m] = per_header

    by_code, sections = {}, {}
    section = None
    for row_no, row in enumerate(all_rows[HEADER_ROW_QUARTER - 1:], start=HEADER_ROW_QUARTER):
        col_a = row[0] if len(row) > 0 else None
        name = row[1] if len(row) > 1 else None
        code = row[code_col] if len(row) > code_col else None

        # A section header has a label in column A and NOTHING in column B.
        # "no Tadawul code" alone is not enough: DOME International (row 489) is
        # a real company whose code cell is a VLOOKUP that returns blank, and on
        # that looser rule it was read as a section - handing its name to every
        # company below it, and with it the wrong set of column labels.
        # Row 3 is the exception: it carries the first section name and a date.
        if col_a and not code and (not name or row_no == HEADER_ROW_QUARTER):
            section = str(col_a).strip()
            sections.setdefault(section, [])
            continue
        if code is None:
            continue
        code = str(code).strip()
        entry = {"row": row_no, "section": section, "name": str(name) if name else ""}
        by_code.setdefault(code, []).append(entry)
        sections.setdefault(section, []).append(code)

    wb.close()

    def half_cols_for_row(row_no, metric):
        """The half-year labels that apply to a given row: the nearest header above it."""
        applicable = [hr for hr in half_header_rows if hr < row_no]
        if not applicable:
            return {}
        return metric_half_cols[metric].get(max(applicable), {})

    def in_scope(code):
        code = str(code).strip()
        if code in KEEP_CODES:
            return True, None
        if code in SKIP_CODES:
            return False, "insurance company (filed under another section in the sheet)"
        for e in by_code.get(code, []):
            if e["section"] and e["section"].strip().lower() in SKIP_SECTIONS:
                return False, f"{e['section']} section is out of scope"
        return True, None

    return {"by_code": by_code, "sections": sections, "in_scope": in_scope,
            "metric_cols": metric_cols, "metric_half_cols": metric_half_cols,
            "half_cols_for_row": half_cols_for_row, "metric_blocks": metric_blocks,
            "half_year_header_rows": half_header_rows,
            "halfyear_section": NOMU_HALFYEAR_SECTION, "code_col": code_col + 1}


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    idx = load_index()
    print(f"workbook : {WORKBOOK.name}  /  sheet '{SHEET}'")
    print(f"codes    : {len(idx['by_code'])} distinct, "
          f"{sum(len(v) for v in idx['by_code'].values())} rows")
    dupes = {c: v for c, v in idx["by_code"].items() if len(v) > 1}
    print(f"duplicates: {len(dupes)} codes on two rows (figures go to both)")
    out = [c for c in idx["by_code"] if not idx["in_scope"](c)[0]]
    print(f"out of scope: {len(out)} codes")
    print("\nsections:")
    for s, codes in idx["sections"].items():
        mark = "  <- skipped" if s and s.strip().lower() in SKIP_SECTIONS else ""
        print(f"   {str(s):<28} {len(codes):>3}{mark}")
    for c in ("8070", "8313"):
        ok, why = idx["in_scope"](c)
        where = idx["by_code"].get(c, [{}])[0]
        print(f"\n   {c} {where.get('name','?')[:28]:<28} section={where.get('section')}  "
              f"-> {'IN SCOPE' if ok else 'SKIPPED: ' + why}")
