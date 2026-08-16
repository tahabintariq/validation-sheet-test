"""
Reads how the "Morning meeting sheet" is wired, so its tables can be rebuilt
without opening Excel.

WHY THIS EXISTS
The intern's job was: type a ticker into A2 (or A24 for the six-month
companies), let the sheet fill in, screenshot it, repeat. Nothing is typed into
that table - every cell is a formula. So the whole thing can be reproduced from
the sheets it reads, and the screenshots stop being necessary.

Verified against Derayah, whose values were cached in the workbook: Revenue,
Operating Income and Net Income matched across all twelve quarters, and Market
Cap matched exactly.

HOW THE SHEET ACTUALLY WORKS
Each figure is:

    INDEX('Qtrly Results Updated'!$H$4:$EE$1059,
          MATCH($A$2,  ...!$A$4:$A$1059, 0),      <- the company's row, by ticker
          MATCH(AI$4,  ...!$H$1:$EE$1,   0))      <- the period's column

The second MATCH does NOT look for a period label. It looks for a NUMBER held in
row 1 of 'Qtrly Results Updated' - the index that sheet uses for its own
VLOOKUPs. The helper cells (AI$4, AI$5, ...) just hold copies of those numbers,
one row per metric. That indirection exists because the numbers are not
sequential (…147, 148, 337, 365, 366…), so no offset arithmetic would work.

NOTHING IS HARDCODED. The metric rows are found by their labels in column C, the
periods by the labels in the header row, and the helper cell for every single
figure is parsed out of that cell's own formula. If someone inserts a column or
moves a row, this follows.
"""

import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

SHEET = "Morning meeting sheet"
SOURCE = "Qtrly Results Updated"
MKTCAP = "Market cap"
UNIVERSE = "Universe"

# Rows read straight from the source sheet. Everything else in the table is
# arithmetic on these, done in stage5b_morning.py.
LOOKUP_LABELS = ["Revenue", "Gross Profit", "Operating Income", "Net Income"]

TICKER = re.compile(r"^[A-Z0-9./ &-]+ AB Equity$", re.I)
HELPER = re.compile(r"MATCH\(\$?([A-Z]{1,3})\$?(\d+),\s*'" + SOURCE + r"'!\$H\$1")
PERIOD = re.compile(r"^([1-4]Q\d{4}|\d{4}Q[1-4]|[12]H\d{4})$", re.I)


def _cell(grid, r, c):
    row = grid[r - 1] if 0 < r <= len(grid) else []
    return row[c - 1] if 0 < c <= len(row) else None


def load_layout(workbook):
    """Work out both blocks of the morning sheet: where the ticker goes, which
    row holds which figure, which columns are which period, and - per figure,
    per period - which helper cell drives the lookup."""
    wbf = openpyxl.load_workbook(workbook, data_only=False, read_only=True)
    frm = [list(r) for r in wbf[SHEET].iter_rows(min_row=1, max_row=60,
                                                 max_col=60, values_only=True)]
    wbf.close()
    wbv = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    val = [list(r) for r in wbv[SHEET].iter_rows(min_row=1, max_row=60,
                                                 max_col=60, values_only=True)]
    wbv.close()

    # --- the blocks are anchored on the cells holding a Bloomberg ticker -----
    anchors = [r for r in range(1, 41)
               if isinstance(_cell(val, r, 1), str) and TICKER.match(_cell(val, r, 1).strip())]
    if len(anchors) < 2:
        raise SystemExit(
            f"FATAL: expected two ticker cells in column A of '{SHEET}', found {anchors}.\n"
            "  The sheet layout has changed - refusing to guess which table is which.")

    blocks = []
    for i, anchor in enumerate(anchors):
        end = anchors[i + 1] - 1 if i + 1 < len(anchors) else anchor + 20

        # metric rows, by their label in column C
        rows = {}
        for r in range(anchor, end + 1):
            lab = _cell(val, r, 3)
            if isinstance(lab, str) and lab.strip():
                rows.setdefault(lab.strip(), r)

        # the period header row: the one with the most period-looking labels
        best, best_n = None, 0
        for r in range(anchor, end + 1):
            n = sum(1 for c in range(3, 20)
                    if isinstance(_cell(val, r, c), str)
                    and PERIOD.match(str(_cell(val, r, c)).strip()))
            if n > best_n:
                best, best_n = r, n
        period_row = best

        periods = [(c, str(_cell(val, period_row, c)).strip())
                   for c in range(3, 20)
                   if isinstance(_cell(val, period_row, c), str)
                   and PERIOD.match(str(_cell(val, period_row, c)).strip())]
        # the "Actual" column is the newest, and appears again under Concensus /
        # Hilal further right. Only the run of distinct periods is the table.
        seen, cols = set(), []
        for c, p in periods:
            if p in seen:
                break
            seen.add(p)
            cols.append((c, p))

        # per figure, per column: the helper cell its own formula points at
        lookups = {}
        for label in LOOKUP_LABELS:
            r = rows.get(label)
            if not r:
                continue
            per_col = {}
            for c, _ in cols:
                f = _cell(frm, r, c)
                m = HELPER.search(f) if isinstance(f, str) else None
                if m:
                    per_col[c] = (column_index_from_string(m.group(1)), int(m.group(2)))
            if per_col:
                lookups[label] = per_col

        # Which columns each line is actually filled in for. The first four
        # columns of every derived row are EMPTY in the sheet - there is no
        # formula there at all, because a year-on-year change needs a year of
        # history behind it. Computing a value where the sheet leaves a blank
        # would put numbers in the email that the analyst has never seen.
        populated = {}
        for label, r in rows.items():
            populated[label] = {c for c, _ in cols
                                if isinstance(_cell(frm, r, c), str)
                                and str(_cell(frm, r, c)).startswith("=")}

        missing = [l for l in LOOKUP_LABELS if l not in lookups]
        if missing:
            raise SystemExit(
                f"FATAL: in the block at row {anchor}, no lookup formula found for "
                f"{missing}.\n  The morning sheet has changed shape - refusing to guess.")

        halfyear = any(p.upper().startswith(("1H", "2H")) for _, p in cols)
        blocks.append({
            "name": "half-year" if halfyear else "quarterly",
            "ticker_row": anchor,
            "name_row": rows.get(next((k for k in rows if k in ("",)), ""), anchor + 2),
            "period_row": period_row,
            "cols": cols,                       # [(column index, period label)]
            "rows": rows,                       # label -> row
            "populated": populated,             # label -> {columns the sheet fills}
            "lookups": lookups,                 # label -> {col: (helper col, helper row)}
            "halfyear": halfyear,
            # A year is 4 quarters or 2 halves. Both the year-on-year comparison
            # and the annualising multiplier follow from that, so neither is
            # written down twice.
            "periods_per_year": 2 if halfyear else 4,
        })

    return {"blocks": blocks, "values": val, "formulas": frm}


def source_tables(workbook):
    """The sheets the morning table reads from."""
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    src = [list(r) for r in wb[SHEET if False else SOURCE].iter_rows(values_only=True)]
    mkt = [list(r) for r in wb[MKTCAP].iter_rows(values_only=True)]
    uni = [list(r) for r in wb[UNIVERSE].iter_rows(values_only=True)]
    wb.close()

    # row 1 of the source sheet holds the index numbers the MATCH looks for
    index_row = src[0] if src else []
    by_index = {}
    for c in range(8, min(136, len(index_row) + 1)):        # H..EE
        v = index_row[c - 1]
        if isinstance(v, (int, float)) and v not in by_index:
            by_index[v] = c

    by_ticker = {}
    for r in range(4, len(src) + 1):
        t = src[r - 1][0] if src[r - 1] else None
        if isinstance(t, str) and t.strip():
            by_ticker.setdefault(t.strip(), r)

    mkt_by_ticker = {}
    for row in mkt:
        t = row[0] if row else None
        if isinstance(t, str) and t.strip() and len(row) > 9:
            mkt_by_ticker.setdefault(t.strip(), row[9])       # column J

    name_by_ticker = {}
    for row in uni:
        t = row[0] if row else None
        if isinstance(t, str) and t.strip() and len(row) > 1:
            name_by_ticker.setdefault(t.strip(), row[1])      # column B

    return {"src": src, "index_to_col": by_index, "row_by_ticker": by_ticker,
            "mktcap": mkt_by_ticker, "name": name_by_ticker}


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    from sheet_index import WORKBOOK
    lay = load_layout(WORKBOOK)
    for b in lay["blocks"]:
        print(f"\n{b['name']} block - ticker in A{b['ticker_row']}, "
              f"periods from row {b['period_row']}")
        print(f"  periods : {[p for _, p in b['cols']]}")
        print(f"  figures : {[(k, v) for k, v in b['rows'].items()][:12]}")
        for label, per in b["lookups"].items():
            first = sorted(per.items())[0]
            print(f"    {label:<18} helper column {get_column_letter(first[1][0])}"
                  f"{first[1][1]} for the first period")
