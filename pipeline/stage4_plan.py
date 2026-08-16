"""
Stage 4a - work out exactly which cell gets which number. WRITES NOTHING.

Produces output/04_write_preview.json plus a readable summary, so the plan can
be checked against the sheet before anything touches the workbook.

Decisions made here:
  - which quarter the announcement covers, taken from its stated period end
    (2026-06-30 -> 2026 Q2). Never inferred from the announcement date, which
    would be wrong for late filings and for corrections issued weeks later.
  - which column that quarter is in, per metric, looked up by header name
  - which row(s) the company is on - both rows when a code appears twice
  - what is already in the target cell, and therefore whether to write

A cell is only written when it is empty, or when a correction supersedes a value
we put there. Anything else - a formula in the way, a missing column, an
unexpected existing number - is reported and left alone.
"""

import json
import sys
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from sheet_index import WORKBOOK, SHEET, load_index

PKT = timezone(timedelta(hours=5))
HERE = Path(__file__).resolve().parent
OUT_DIR = HERE / "output"

METRIC_KEYS = [("revenue", "Revenue"), ("gross_profit", "Gross Profit"),
               ("operating_profit", "Operating Profit"), ("net_income", "Net Income")]

NOMU_HALFYEAR = "nomu halfyear companies"


def quarter_from_period_end(period_end):
    """'2026-06-30' -> (2026, 2). None if the date is not a quarter end."""
    if not period_end:
        return None
    try:
        y, m, d = (int(x) for x in period_end.split("-"))
    except ValueError:
        return None
    return (y, {3: 1, 6: 2, 9: 3, 12: 4}.get(m)) if m in (3, 6, 9, 12) else None


def main():
    src = json.load(open(OUT_DIR / "03_extracted.json", encoding="utf-8"))
    idx = load_index()

    # data_only=False so we see FORMULAS as formulas. A cell holding a VLOOKUP
    # must never be overwritten - that would sever the company's link to the feed.
    wb = openpyxl.load_workbook(WORKBOOK, data_only=False)
    ws = wb[SHEET]

    updated_on_col = None
    for i, c in enumerate(ws[2], start=1):
        if c.value and str(c.value).strip().lower() == "updated on":
            updated_on_col = i
            break

    plan, skipped = [], []
    for e in src["extractions"]:
        code = e["tadawul_code"]
        rows = idx["by_code"].get(code, [])
        # The announcement may carry both a quarter and a six-month set. Which one
        # belongs in the sheet depends on the ROW, so the choice is made below,
        # once we know which section the row sits in.
        qm = e.get("quarter_metrics") or {}
        pm = e.get("period_metrics") or {}
        cm = e.get("correction_metrics") or {}
        has_figures = any(b.get(k) for b in (qm, pm, cm) for k, _ in METRIC_KEYS)

        if not has_figures:
            skipped.append({**{k: e[k] for k in ("announcement_id", "tadawul_code", "company")},
                            "reason": e.get("note") or "no figures in this announcement"})
            continue
        if e.get("needs_review"):
            # Stage 3 was not confident. A figure nobody has checked must not reach
            # the sheet just because a cell happened to be empty.
            skipped.append({**{k: e[k] for k in ("announcement_id", "tadawul_code", "company")},
                            "reason": "flagged for review in Stage 3 - " + (e.get("note") or "")})
            continue
        if not rows:
            skipped.append({**{k: e[k] for k in ("announcement_id", "tadawul_code", "company")},
                            "reason": "company is not in the sheet - report in the email"})
            continue

        q = quarter_from_period_end(e.get("period_end"))
        if not q:
            skipped.append({**{k: e[k] for k in ("announcement_id", "tadawul_code", "company")},
                            "reason": f"could not tell which quarter from period end {e.get('period_end')!r}"})
            continue

        # Half-year reporters use the section's own header row, which relabels the
        # same columns as 1H/2H periods. A June period end is the first half; a
        # December one is the second. The sheet's own labels decide the column -
        # not the quarter, which for these rows means something else entirely.
        half = (q[0], 1) if q[1] == 2 else ((q[0], 2) if q[1] == 4 else None)

        for entry in rows:                       # both rows when a code is duplicated
            # Decided per ROW, not per company: the same company can sit in a
            # quarterly section on one row and a half-year section on another
            # (QOMEL is in both "NOMU Halfyear Companies" and "IPOs - 2024
            # Halfyearly"), and the two use different columns for the same period.
            section = (entry["section"] or "").strip().lower()
            is_halfyear = "halfyear" in section.replace(" ", "")
            if is_halfyear and half is None:
                skipped.append({**{k: e[k] for k in ("announcement_id", "tadawul_code", "company")},
                                "reason": f"half-year reporter but period ends in Q{q[1]} - expected June or December"})
                continue

            # quarterly rows take the quarter figures; half-year rows take the
            # cumulative six-month figures. Getting this backwards would put a
            # three-month number in a six-month cell - about half the true value,
            # and completely plausible-looking in the sheet.
            # An annual filing arrives already reduced to the period the sheet
            # wants: Stage 3 puts the derived Q4 in quarter_metrics and the
            # derived second half in period_metrics. So the routing below is
            # unchanged - a December period end simply resolves to Q4 / 2H.
            metrics = cm or (pm if is_halfyear else qm)
            basis = "correction" if cm else ("period (six months)" if is_halfyear else "quarter")
            if metrics and any((metrics.get(k) or {}).get("derivation")
                               for k, _ in METRIC_KEYS):
                basis = ("second half (full year minus 1H)" if is_halfyear
                         else "Q4 (full year minus Q1-Q3)")
            if not metrics:
                skipped.append({**{k: e[k] for k in ("announcement_id", "tadawul_code", "company")},
                                "reason": f"row needs the {basis} figures but the announcement does not report them"})
                continue

            for key, header in METRIC_KEYS:
                m = metrics.get(key)
                if not m or m.get("millions") is None:
                    continue
                if is_halfyear:
                    period_label = f"{half[1]}H{half[0]}"
                    col = idx["half_cols_for_row"](entry["row"], header).get(half)
                else:
                    period_label = f"{q[1]}Q{q[0]}"
                    col = idx["metric_cols"][header].get(q)
                if not col:
                    plan.append({"announcement_id": e["announcement_id"], "tadawul_code": code,
                                 "company": e["company"], "row": entry["row"], "metric": header,
                                 "period": period_label, "cell": None, "value": m["millions"],
                                 "action": "blocked", "blocked_reason": "missing_column",
                                 "note": f"no {period_label} column for {header} - add it in Excel"})
                    continue

                cell = ws.cell(row=entry["row"], column=col)
                addr = f"{get_column_letter(col)}{entry['row']}"
                existing = cell.value
                is_formula = isinstance(existing, str) and existing.startswith("=")

                blocked_reason = None
                if is_formula:
                    action, note = "blocked", "cell holds a formula (feed link) - left alone, reported in the email"
                    blocked_reason = "formula_in_cell"
                elif existing is None or existing == "":
                    action, note = "write", ""
                elif e.get("page_type") == "correction" or e.get("corrects_a_number"):
                    action, note = "overwrite", f"correction supersedes {existing}"
                else:
                    action, note = "blocked", f"cell already has {existing} - left alone"
                    blocked_reason = "cell_has_value"

                plan.append({
                    "announcement_id": e["announcement_id"], "tadawul_code": code,
                    "company": e["company"], "row": entry["row"], "section": entry["section"],
                    "metric": header, "period": period_label, "cell": addr,
                    "value": m["millions"], "source_line": (m.get("lines") or [None])[0],
                    "as_printed": m.get("as_printed"), "units": e.get("units_as_printed"), "basis": basis,
                    # kept so the daily email can show the working behind a Q4:
                    # which full-year figure, less which cells.
                    "derivation": m.get("derivation"),
                    "existing": existing if not is_formula else "(formula)",
                    "action": action, "note": note, "blocked_reason": blocked_reason,
                })

    # --- two announcements landing on the same cell --------------------------
    # A company's results and a later correction to them both target the same
    # cell. Without this the plan would hold both and the last one written would
    # silently win. The newer announcement supersedes the older.
    by_cell = {}
    for p in plan:
        if p["cell"] and p["action"] in ("write", "overwrite"):
            by_cell.setdefault(p["cell"], []).append(p)
    conflicts = []
    for addr, entries in by_cell.items():
        if len(entries) < 2:
            continue
        entries.sort(key=lambda x: x["announcement_id"], reverse=True)
        winner, losers = entries[0], entries[1:]
        winner["action"] = "overwrite"
        winner["note"] = (f"supersedes announcement(s) "
                          f"{', '.join(str(l['announcement_id']) for l in losers)} on this cell")
        for l in losers:
            l["action"] = "superseded"
            l["note"] = f"superseded by announcement {winner['announcement_id']}"
        conflicts.append({"cell": addr, "company": winner["company"], "metric": winner["metric"],
                          "winner": {"announcement_id": winner["announcement_id"], "value": winner["value"],
                                     "source_line": winner.get("source_line")},
                          "superseded": [{"announcement_id": l["announcement_id"], "value": l["value"],
                                          "source_line": l.get("source_line")} for l in losers]})

    # --- stamp "Updated On" only on rows that actually received a figure ------
    if updated_on_col:
        stamped = {}
        for p in plan:
            if p["action"] in ("write", "overwrite") and p["metric"] != "Updated On":
                stamped.setdefault(p["row"], p)
        for row_no, p in stamped.items():
            plan.append({"announcement_id": p["announcement_id"], "tadawul_code": p["tadawul_code"],
                         "company": p["company"], "row": row_no, "metric": "Updated On",
                         "period": p["period"],
                         "cell": f"{get_column_letter(updated_on_col)}{row_no}",
                         "value": datetime.now(PKT).strftime("%Y-%m-%d"),
                         "action": "write", "note": "run date"})

    wb.close()

    counts = Counter(p["action"] for p in plan)
    artifact = {
        "_stage": "04_write_preview",
        "_note": "Planned cell writes. NOTHING has been written. Each entry names the "
                 "exact cell, the value, and which line on the announcement it came from.",
        "planned_at_pkt": datetime.now(PKT).strftime("%d/%m/%Y %H:%M:%S"),
        "workbook": str(WORKBOOK), "sheet": SHEET,
        "counts": dict(counts),
        "cells_to_write": counts["write"] + counts["overwrite"],
        "companies": len({p["tadawul_code"] for p in plan if p["action"] in ("write", "overwrite")}),
        "conflicts": conflicts,
        "skipped": skipped,
        "plan": plan,
    }
    out = OUT_DIR / "04_write_preview.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(artifact, f, indent=2, ensure_ascii=False)

    print(f"Stage 4a - write plan (nothing written)\n")
    print(f"  workbook : {WORKBOOK.name}  /  '{SHEET}'")
    print(f"  actions  : " + "   ".join(f"{k} {v}" for k, v in counts.items()))
    print(f"  skipped  : {len(skipped)} announcements\n")
    shown = [p for p in plan if p["action"] in ("write", "overwrite") and p["metric"] != "Updated On"]
    print(f"  {'cell':<8} {'company':<20} {'metric':<17} {'value':>12}   from")
    for p in shown[:14]:
        print(f"  {p['cell']:<8} {p['company'][:19]:<20} {p['metric']:<17} {p['value']:>12}   {str(p.get('source_line'))[:34]}")
    print(f"  ... {len(shown)-14} more" if len(shown) > 14 else "")
    blocked = [p for p in plan if p["action"] == "blocked"]
    if blocked:
        print(f"\n  BLOCKED ({len(blocked)}):")
        for p in blocked[:8]:
            print(f"   {str(p['cell']):<8} {p['company'][:18]:<18} {p['metric']:<17} {p['note']}")
    print(f"\n  written : {out}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    main()
